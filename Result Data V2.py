from __future__ import annotations

import argparse
import os
import re
import sys
import time
import webbrowser
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.utils.datetime import from_excel

# Optional import for the education-board browser verification page
try:
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PlaywrightTimeoutError = TimeoutError
    PLAYWRIGHT_AVAILABLE = False

# Optional imports for Gemini Auto-Captcha
try:
    from google import genai
    from PIL import Image
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False


# ==============================================================================
# CONFIGURATION
# ==============================================================================
# Optional default Gemini API key. Prefer --gemini-api-key or GEMINI_API_KEY.
GEMINI_API_KEY = ""

# The chosen model based on optimal RPM (15) and RPD (500) limits
GEMINI_MODEL = "gemini-3.1-flash-lite"
# ==============================================================================


VERSION = "2.2.0"
WORKBOOK_NAME = "Result data.xlsx"

BASE_URL = "https://www.educationboardresults.gov.bd"
HOME_URL = f"{BASE_URL}/v2/home"
CAPTCHA_URL = f"{BASE_URL}/v2/captcha"
RESULT_URL = f"{BASE_URL}/v2/getres"
CHALLENGE_PATH = "/_challenge-engine"

BTEB_BASE_URL = "https://result.bteb.gov.bd"
BTEB_SEARCH_URL = f"{BTEB_BASE_URL}/result-search"
BTEB_PUBLIC_RESULT_URL = f"{BTEB_BASE_URL}/api/public/result"

INPUT_COLUMNS = ["Exam", "Passing Year", "Board", "Roll", "Reg"]
OUTPUT_COLUMNS = ["Name", "Father's Name", "Mother's Name", "GPA", "Date of Birth"]

ERROR_STATUSES = {
    "Result Not Found",
    "Invalid Criteria",
    "HTTP Error",
    "Extraction Error",
    "Missing Input Data",
    "Captcha Not Found",
    "Captcha Parse Error",
    "Unknown Exam Type",
    "Unknown Board",
    "HTML Parse Error (Captcha)",
    "Unexpected Error during Captcha",
    "Extraction Failed",
    "Manual Captcha Skipped",
    "Captcha Failed",
    "Result Error",
    "Missing Technical Input",
    "Unknown Technical Exam",
    "Technical Result Error",
}

RAW_EXAM_TYPE_MAPPING = {
    "JSC/JDC": "jsc",
    "JSC or Equivalent": "jsc",
    "SSC/Dakhil/Equivalent": "ssc",
    "SSC/Dakhil": "ssc",
    "SSC or Equivalent": "ssc",
    "SSC(Vocational)": "ssc",
    "HSC/Alim/Equivalent": "hsc",
    "HSC/Alim": "hsc",
    "HSC(Vocational)": "hsc",
    "HSC(BM)": "hsc",
    "Diploma in Commerce": "hsc",
    "DIBS (Diploma in Business Studies)": "dibs",
    "Diploma in Business Studies": "dibs",
    "DIBS": "dibs",
}

RAW_BOARD_MAPPING = {
    "Barisal": "barisal",
    "Chittagong": "chittagong",
    "Comilla": "comilla",
    "Dhaka": "dhaka",
    "Dinajpur": "dinajpur",
    "Jessore": "jessore",
    "Mymensingh": "mymensingh",
    "Rajshahi": "rajshahi",
    "Sylhet": "sylhet",
    "Madrasah": "madrasah",
    "Technical": "tec",
    "TEC": "tec",
}

OPTIONAL_BTEB_COLUMNS = {
    "curriculum": [
        "Curriculum",
        "Curriculum Code",
        "BTEB Curriculum",
        "BTEB Curriculum Code",
        "Technical Curriculum",
    ],
    "semester": [
        "Semester",
        "Class",
        "Semester / Class",
        "BTEB Semester",
        "Technical Semester",
    ],
}


@dataclass
class StudentResult:
    name: str
    father_name: str
    mother_name: str
    gpa: str
    date_of_birth: str


class LookupErrorWithStatus(Exception):
    def __init__(self, status: str, message: str):
        super().__init__(message)
        self.status = status
        self.message = message


def normalize_key(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


EXAM_TYPE_MAPPING = {normalize_key(key): value for key, value in RAW_EXAM_TYPE_MAPPING.items()}
BOARD_MAPPING = {normalize_key(key): value for key, value in RAW_BOARD_MAPPING.items()}


def excel_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_date_of_birth(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")

    if isinstance(value, (int, float)):
        try:
            return from_excel(value).strftime("%d-%m-%Y")
        except Exception:
            return excel_text(value)

    text = excel_text(value)
    if not text or text.upper() == "N/A":
        return text or "N/A"

    text = text.split()[0].strip()

    iso_match = re.fullmatch(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if iso_match:
        year, month, day = [int(part) for part in iso_match.groups()]
        try:
            return date(year, month, day).strftime("%d-%m-%Y")
        except ValueError:
            return text

    bd_match = re.fullmatch(r"(\d{1,2})[-/.](\d{1,2})[-/.](\d{2,4})", text)
    if bd_match:
        day, month, year = [int(part) for part in bd_match.groups()]
        if year < 100:
            year += 2000 if year <= 30 else 1900
        try:
            return date(year, month, day).strftime("%d-%m-%Y")
        except ValueError:
            return text

    return text


def resolve_workbook_path(path_arg: str | None) -> Path:
    if path_arg:
        return Path(path_arg).expanduser().resolve()

    cwd_candidate = Path.cwd() / WORKBOOK_NAME
    if cwd_candidate.exists():
        return cwd_candidate.resolve()

    script_candidate = Path(__file__).resolve().parent / WORKBOOK_NAME
    if script_candidate.exists():
        return script_candidate.resolve()

    return cwd_candidate.resolve()


def build_column_map(sheet: Any) -> dict[str, int]:
    column_map: dict[str, int] = {}
    for index, cell in enumerate(sheet[1], start=1):
        header = excel_text(cell.value)
        if header:
            column_map[header] = index
    return column_map


def require_columns(column_map: dict[str, int]) -> None:
    missing = [column for column in INPUT_COLUMNS + OUTPUT_COLUMNS if column not in column_map]
    if missing:
        joined = ", ".join(missing)
        raise RuntimeError(f"Required column(s) not found in the Excel sheet: {joined}")


def create_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "en-US,en;q=0.9",
            "Referer": HOME_URL,
        }
    )
    return session


def is_challenge_response(response: requests.Response) -> bool:
    url = response.url.lower()
    if CHALLENGE_PATH in url:
        return True

    content_type = response.headers.get("Content-Type", "").lower()
    if "text/html" not in content_type:
        return False

    text = response.text.lower()
    return "verifying your browser" in text or CHALLENGE_PATH in text


def is_image_response(response: requests.Response) -> bool:
    content_type = response.headers.get("Content-Type", "").lower()
    return response.ok and "image" in content_type


def launch_verification_browser(playwright: Any, headless: bool) -> Any:
    launch_options = [
        {"channel": "msedge", "headless": headless},
        {"channel": "chrome", "headless": headless},
        {"headless": headless},
    ]

    last_error: Exception | None = None
    for options in launch_options:
        try:
            return playwright.chromium.launch(**options)
        except Exception as exc:
            last_error = exc

    raise RuntimeError(f"Could not launch a Chromium browser: {last_error}")


def page_has_result_form(page: Any) -> bool:
    try:
        return bool(page.locator("#form, #captcha_img, select#board, select#exam").count())
    except Exception:
        return False


def copy_browser_cookies_to_session(session: requests.Session, cookies: list[dict[str, Any]]) -> None:
    for cookie in cookies:
        name = cookie.get("name")
        value = cookie.get("value")
        if not name or value is None:
            continue
        session.cookies.set(
            name,
            value,
            domain=cookie.get("domain"),
            path=cookie.get("path", "/"),
        )


def verify_education_board_session_with_browser(
    session: requests.Session,
    *,
    timeout_seconds: int,
    headless: bool,
) -> None:
    if not PLAYWRIGHT_AVAILABLE:
        raise LookupErrorWithStatus(
            "Browser Verification Required",
            "The education-board site returned its browser verification page. "
            "Install Playwright with 'pip install playwright' and run "
            "'python -m playwright install chromium', then run the script again.",
        )

    timeout_ms = max(5, timeout_seconds) * 1000
    print("Education-board browser verification required. Opening a browser session...")

    try:
        with sync_playwright() as playwright:
            browser = None
            context = None
            browser = launch_verification_browser(playwright, headless=headless)
            try:
                context = browser.new_context()
                page = context.new_page()
                page.goto(HOME_URL, wait_until="domcontentloaded", timeout=timeout_ms)

                deadline = time.monotonic() + timeout_seconds
                while time.monotonic() < deadline:
                    if CHALLENGE_PATH not in page.url and page_has_result_form(page):
                        break
                    page.wait_for_timeout(500)

                if not page_has_result_form(page):
                    if headless:
                        raise LookupErrorWithStatus(
                            "Browser Verification Required",
                            "The headless browser did not reach the result form. "
                            "Run again without --headless-browser-verify so the verification window can be shown.",
                        )
                    input(
                        "Complete the browser verification if needed. "
                        "When the result form is visible, press Enter here..."
                    )

                copy_browser_cookies_to_session(session, context.cookies(BASE_URL))
            finally:
                if context:
                    context.close()
                if browser:
                    browser.close()
    except PlaywrightTimeoutError as exc:
        raise LookupErrorWithStatus(
            "Browser Verification Required",
            f"The browser verification page did not finish within {timeout_seconds} seconds: {exc}",
        ) from exc
    except RuntimeError as exc:
        raise LookupErrorWithStatus("Browser Verification Required", str(exc)) from exc

    print("Browser verification complete. Continuing with the regular CAPTCHA flow.")


def create_bteb_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.9",
            "Content-Type": "application/json",
            "Origin": BTEB_BASE_URL,
            "Referer": BTEB_SEARCH_URL,
        }
    )
    return session


def find_first_column(column_map: dict[str, int], candidates: list[str]) -> int | None:
    for candidate in candidates:
        if candidate in column_map:
            return column_map[candidate]
    return None


def get_optional_cell(sheet: Any, row_idx: int, column_map: dict[str, int], key: str) -> str:
    column = find_first_column(column_map, OPTIONAL_BTEB_COLUMNS[key])
    if not column:
        return ""
    return excel_text(sheet.cell(row=row_idx, column=column).value)


def is_technical_board(board_text: str, mapped_board: str | None = None) -> bool:
    normalized = normalize_key(board_text)
    return normalized in {"technical", "tec"} or mapped_board == "tec"


def normalize_bteb_code(value: str) -> str:
    match = re.search(r"\d+", value or "")
    return match.group(0) if match else ""


def normalize_bteb_semester(value: str) -> str:
    normalized = normalize_key(value)
    if "class 9" in normalized or normalized == "9":
        return "1"
    if "class 10" in normalized or normalized == "10":
        return "2"
    return normalize_bteb_code(value)


def infer_bteb_curriculum(exam_text: str, override: str = "") -> str:
    override_code = normalize_bteb_code(override)
    if override_code:
        return override_code

    normalized = normalize_key(exam_text)
    has_vocational = "voc" in normalized or "vocation" in normalized or "technical" in normalized

    if "dakhil" in normalized and "ssc" not in normalized and has_vocational:
        return "77"
    if "ssc" in normalized or "dakhil" in normalized:
        return "27"
    if "hsc" in normalized and ("voc" in normalized or "bm" in normalized or "business management" in normalized):
        return "26"
    if "diploma in commerce" in normalized:
        return "25"
    if "textile" in normalized:
        return "19"

    raise LookupErrorWithStatus(
        "Unknown Technical Exam",
        f"Could not infer BTEB curriculum from exam value '{exam_text}'. "
        "Add a 'Curriculum Code' column, for example 27 for SSC (Vocational).",
    )


def infer_bteb_semester(exam_text: str, override: str = "") -> str:
    override_code = normalize_bteb_semester(override)
    if override_code:
        return override_code

    normalized = normalize_key(exam_text)
    if "class 9" in normalized:
        return "1"
    if "class 10" in normalized or "ssc" in normalized or "dakhil" in normalized:
        return "2"
    if "hsc" in normalized and ("voc" in normalized or "bm" in normalized or "business management" in normalized):
        return "2"

    raise LookupErrorWithStatus(
        "Missing Technical Input",
        f"Could not infer BTEB semester/class from exam value '{exam_text}'. "
        "Add a 'Semester' or 'Class' column, for example 2 for Class 10.",
    )


def open_captcha_image(path: Path) -> None:
    try:
        if os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
        else:
            webbrowser.open(path.as_uri())
    except Exception as exc:
        print(f"Could not open CAPTCHA image automatically: {exc}")


def fetch_captcha(
    session: requests.Session,
    captcha_dir: Path,
    row_idx: int,
    *,
    allow_browser_verify: bool,
    browser_verify_timeout: int,
    headless_browser_verify: bool,
) -> Path:
    home_response = session.get(HOME_URL, timeout=60)
    if is_challenge_response(home_response):
        if not allow_browser_verify:
            raise LookupErrorWithStatus(
                "Browser Verification Required",
                "The education-board site returned its browser verification page.",
            )
        verify_education_board_session_with_browser(
            session,
            timeout_seconds=browser_verify_timeout,
            headless=headless_browser_verify,
        )
    else:
        home_response.raise_for_status()

    response = session.get(
        CAPTCHA_URL,
        params={"t": str(int(time.time() * 1000))},
        headers={"Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8"},
        timeout=60,
    )

    if is_challenge_response(response):
        if not allow_browser_verify:
            raise LookupErrorWithStatus(
                "Browser Verification Required",
                "The CAPTCHA endpoint returned the browser verification page.",
            )
        verify_education_board_session_with_browser(
            session,
            timeout_seconds=browser_verify_timeout,
            headless=headless_browser_verify,
        )
        response = session.get(
            CAPTCHA_URL,
            params={"t": str(int(time.time() * 1000))},
            headers={"Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8"},
            timeout=60,
        )

    response.raise_for_status()

    if not is_image_response(response):
        content_type = response.headers.get("Content-Type", "unknown")
        raise LookupErrorWithStatus(
            "Captcha Not Found",
            f"The CAPTCHA endpoint did not return an image. Content-Type was '{content_type}'.",
        )

    captcha_dir.mkdir(parents=True, exist_ok=True)
    captcha_path = captcha_dir / f"captcha_row_{row_idx}.png"
    captcha_path.write_bytes(response.content)
    return captcha_path


def prompt_for_captcha(row_idx: int, captcha_path: Path, should_open: bool) -> str:
    print(f"CAPTCHA for row {row_idx} saved to: {captcha_path}")
    if should_open:
        open_captcha_image(captcha_path)

    captcha = input(f"Enter CAPTCHA digits for row {row_idx} (blank skips this row): ").strip()
    if not captcha:
        raise LookupErrorWithStatus("Manual Captcha Skipped", "No CAPTCHA value entered.")
    return captcha


def solve_captcha_with_gemini(captcha_path: Path, api_key: str) -> str:
    """Uses the new google-genai model to read the text off the CAPTCHA image."""
    client = genai.Client(api_key=api_key)
    img = Image.open(captcha_path)
    prompt = (
        "You are an automated OCR system. Read the exact alphanumeric characters "
        "shown in this CAPTCHA image. Return ONLY the extracted text. "
        "Do not include any spaces, punctuation, markdown formatting, or additional words."
    )
    
    response = client.models.generate_content(
        model=GEMINI_MODEL,
        contents=[img, prompt]
    )
    
    # Clean the response to ensure no weird artifacts
    result = response.text.strip().replace(" ", "").replace("\n", "")
    return result


def extract_gpa(value: Any) -> str:
    text = excel_text(value)
    if not text:
        return "N/A"

    match = re.search(r"GPA\s*=?\s*([0-9.]+)", text, flags=re.IGNORECASE)
    if match:
        return match.group(1)
    return text


def parse_json_result(payload: dict[str, Any]) -> StudentResult:
    status = str(payload.get("status", ""))
    if status != "0":
        message = excel_text(payload.get("msg")) or "Result lookup failed."
        if "captcha" in message.lower():
            raise LookupErrorWithStatus("Captcha Failed", message)
        raise LookupErrorWithStatus("Result Error", message)

    result = payload.get("res")
    if not isinstance(result, dict):
        raise LookupErrorWithStatus("Extraction Failed", "The response did not contain a result object.")

    return StudentResult(
        name=excel_text(result.get("name")) or "N/A",
        father_name=excel_text(result.get("fname")) or "N/A",
        mother_name=excel_text(result.get("mname")) or "N/A",
        gpa=extract_gpa(result.get("gpa") or result.get("res_detail") or result.get("result")),
        date_of_birth=normalize_date_of_birth(result.get("dob")) or "N/A",
    )


def table_pairs_from_html(html: str) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    pairs: dict[str, str] = {}

    for row in soup.select("tr"):
        cells = [cell.get_text(" ", strip=True) for cell in row.find_all(["td", "th"])]
        if len(cells) >= 2:
            pairs[normalize_key(cells[0])] = cells[1]
        if len(cells) >= 4:
            pairs[normalize_key(cells[2])] = cells[3]

    return pairs


def parse_html_result(html: str) -> StudentResult:
    pairs = table_pairs_from_html(html)
    if not pairs:
        raise LookupErrorWithStatus("Extraction Failed", "No result table was found in the HTML response.")

    name = pairs.get("name of student") or pairs.get("name") or "N/A"
    father_name = pairs.get("father's name") or "N/A"
    mother_name = pairs.get("mother's name") or "N/A"
    gpa = pairs.get("gpa") or extract_gpa(pairs.get("result"))
    date_of_birth = normalize_date_of_birth(pairs.get("date of birth")) or "N/A"

    if all(value == "N/A" for value in [name, father_name, mother_name, gpa, date_of_birth]):
        raise LookupErrorWithStatus("Extraction Failed", "The result table did not contain expected fields.")

    return StudentResult(
        name=name,
        father_name=father_name,
        mother_name=mother_name,
        gpa=gpa,
        date_of_birth=date_of_birth,
    )


def parse_bteb_result_payload(payload: dict[str, Any]) -> StudentResult:
    data = payload.get("data")
    if not isinstance(data, dict):
        raise LookupErrorWithStatus("Technical Result Error", "The BTEB response did not contain result data.")

    semesters = data.get("semesters")
    student = data.get("student")
    if not isinstance(student, dict) or not isinstance(semesters, list) or not semesters:
        raise LookupErrorWithStatus("Result Not Found", "No BTEB result found for the given criteria.")

    semester = semesters[0] if isinstance(semesters[0], dict) else {}
    gpa_info = semester.get("gpa") if isinstance(semester.get("gpa"), dict) else {}
    dob_info = student.get("dateOfBirth") if isinstance(student.get("dateOfBirth"), dict) else {}

    return StudentResult(
        name=excel_text(student.get("studentName")) or "N/A",
        father_name=excel_text(student.get("fatherName")) or "N/A",
        mother_name=excel_text(student.get("motherName")) or "N/A",
        gpa=excel_text(gpa_info.get("gpaWithOptional") or gpa_info.get("gpa")) or "N/A",
        date_of_birth=normalize_date_of_birth(dob_info.get("numeric") or student.get("dateOfBirth")) or "N/A",
    )


def submit_bteb_result_request(
    session: requests.Session,
    *,
    curriculum_code: str,
    semester: str,
    year: str,
    roll: str,
    reg: str,
) -> StudentResult:
    payload = {
        "curriculumCode": curriculum_code,
        "rollNo": roll,
        "regNo": reg,
        "semester": semester,
        "examYear": year,
    }

    response = session.post(BTEB_PUBLIC_RESULT_URL, json=payload, timeout=60)
    response.raise_for_status()

    try:
        return parse_bteb_result_payload(response.json())
    except ValueError:
        return parse_html_result(response.text)


def submit_result_request(
    session: requests.Session,
    *,
    board: str,
    exam: str,
    year: str,
    roll: str,
    reg: str,
    captcha: str,
) -> StudentResult:
    payload = {
        "board": board,
        "exam": exam,
        "year": year,
        "result_type": "1",
        "roll": roll,
        "reg": reg,
        "captcha": captcha,
    }

    response = session.post(
        RESULT_URL,
        data=payload,
        headers={
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Origin": BASE_URL,
            "Referer": HOME_URL,
            "X-Requested-With": "XMLHttpRequest",
        },
        timeout=60,
    )
    response.raise_for_status()

    try:
        return parse_json_result(response.json())
    except ValueError:
        return parse_html_result(response.text)


def lookup_result(
    session: requests.Session,
    *,
    board: str,
    exam: str,
    year: str,
    roll: str,
    reg: str,
    row_idx: int,
    captcha_dir: Path,
    open_captcha: bool,
    max_attempts: int,
    use_api: bool,
    api_key: str | None,
    allow_browser_verify: bool,
    browser_verify_timeout: int,
    headless_browser_verify: bool,
) -> StudentResult:
    last_error: LookupErrorWithStatus | None = None

    for attempt in range(1, max_attempts + 1):
        captcha_path = fetch_captcha(
            session,
            captcha_dir,
            row_idx,
            allow_browser_verify=allow_browser_verify,
            browser_verify_timeout=browser_verify_timeout,
            headless_browser_verify=headless_browser_verify,
        )
        captcha = None

        if use_api and api_key:
            try:
                print(f"  [Attempt {attempt}] Automatically solving CAPTCHA with Gemini API ({GEMINI_MODEL})...")
                captcha = solve_captcha_with_gemini(captcha_path, api_key)
                print(f"  [Attempt {attempt}] Gemini solved CAPTCHA as: '{captcha}'")
            except Exception as e:
                print(f"  [Attempt {attempt}] Gemini auto-solve failed: {e}")
                # Fallback to manual if API fails
                captcha = prompt_for_captcha(row_idx, captcha_path, open_captcha)
        else:
            captcha = prompt_for_captcha(row_idx, captcha_path, open_captcha)

        if not captcha:
            raise LookupErrorWithStatus("Captcha Failed", "No CAPTCHA value obtained.")

        try:
            return submit_result_request(
                session,
                board=board,
                exam=exam,
                year=year,
                roll=roll,
                reg=reg,
                captcha=captcha,
            )
        except LookupErrorWithStatus as exc:
            last_error = exc
            print(f"  [Attempt {attempt}/{max_attempts}] Lookup failed for row {row_idx}: {exc.message}")
            if exc.status != "Captcha Failed":
                raise
            # If CAPTCHA fails, wait a short moment before trying again to avoid hammering the endpoint
            if attempt < max_attempts:
               time.sleep(1)

    if last_error:
        raise last_error
    raise LookupErrorWithStatus("Captcha Failed", "CAPTCHA attempts were exhausted.")


def write_result(sheet: Any, row_idx: int, column_map: dict[str, int], result: StudentResult) -> None:
    sheet.cell(row=row_idx, column=column_map["Name"], value=result.name)
    sheet.cell(row=row_idx, column=column_map["Father's Name"], value=result.father_name)
    sheet.cell(row=row_idx, column=column_map["Mother's Name"], value=result.mother_name)
    sheet.cell(row=row_idx, column=column_map["GPA"], value=result.gpa)
    dob_cell = sheet.cell(row=row_idx, column=column_map["Date of Birth"])
    dob_cell.number_format = "@"
    dob_cell.value = normalize_date_of_birth(result.date_of_birth) or "N/A"


def write_error(sheet: Any, row_idx: int, column_map: dict[str, int], status: str) -> None:
    sheet.cell(row=row_idx, column=column_map["Name"], value=status)
    for column in OUTPUT_COLUMNS[1:]:
        sheet.cell(row=row_idx, column=column_map[column], value=None)


def normalize_existing_dob_cells(sheet: Any, column_map: dict[str, int]) -> None:
    dob_col = column_map["Date of Birth"]
    for row_idx in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row_idx, column=dob_col)
        value = cell.value
        if value is None or excel_text(value) == "":
            continue
        cell.number_format = "@"
        cell.value = normalize_date_of_birth(value)


def save_workbook_progress(
    workbook: Any,
    workbook_path: Path,
    sheet: Any,
    column_map: dict[str, int],
) -> bool:
    normalize_existing_dob_cells(sheet, column_map)
    try:
        workbook.save(workbook_path)
        return True
    except PermissionError:
        print(f"\nError saving Excel file. Close '{workbook_path.name}' if it is open, then run again.")
        return False
    except Exception as exc:
        print(f"\nError saving Excel file: {exc}")
        return False


def consume_enter_keypress() -> bool:
    if os.name != "nt":
        return False

    try:
        import msvcrt
    except ImportError:
        return False

    pressed = False
    try:
        while msvcrt.kbhit():
            key = msvcrt.getwch()
            if key in ("\r", "\n"):
                pressed = True
        return pressed
    except OSError:
        return False


def pause_if_enter_requested(
    args: argparse.Namespace,
    workbook: Any,
    workbook_path: Path,
    sheet: Any,
    column_map: dict[str, int],
) -> bool:
    if args.no_enter_pause or not consume_enter_keypress():
        return True

    print("\nPause requested. Saving current progress before pausing...")
    if not save_workbook_progress(workbook, workbook_path, sheet, column_map):
        return False

    try:
        input("Paused. Press Enter to resume...")
    except EOFError:
        pass
    return True


def save_progress_and_handle_pause(
    args: argparse.Namespace,
    workbook: Any,
    workbook_path: Path,
    sheet: Any,
    column_map: dict[str, int],
) -> bool:
    if not save_workbook_progress(workbook, workbook_path, sheet, column_map):
        return False
    return pause_if_enter_requested(args, workbook, workbook_path, sheet, column_map)


def should_skip_row(sheet: Any, row_idx: int, column_map: dict[str, int], force: bool) -> bool:
    if force:
        return False

    current_name = excel_text(sheet.cell(row=row_idx, column=column_map["Name"]).value)
    return bool(current_name)


def extract_results_to_excel(args: argparse.Namespace) -> int:
    workbook_path = resolve_workbook_path(args.file)
    if not workbook_path.exists():
        print(f"Error: Excel file not found at '{workbook_path}'")
        return 1

    try:
        workbook = openpyxl.load_workbook(workbook_path)
        sheet = workbook.active
    except Exception as exc:
        print(f"Error loading Excel file: {exc}")
        return 1

    try:
        column_map = build_column_map(sheet)
        require_columns(column_map)
    except RuntimeError as exc:
        print(f"Error: {exc}")
        return 1

    # Configure Gemini API using the new google-genai approach
    api_key = args.gemini_api_key or os.environ.get("GEMINI_API_KEY") or GEMINI_API_KEY
    use_api = False
    
    if api_key:
        if not GEMINI_AVAILABLE:
            print("Error: Gemini API Key provided, but 'google-genai' or 'Pillow' is not installed.")
            print("Run: pip install -U google-genai pillow")
            return 1
        use_api = True
        print(f"Gemini API configured (Model: {GEMINI_MODEL}). Automated CAPTCHA solving is ENABLED.")
    else:
        print("No Gemini API key provided. Operating in MANUAL CAPTCHA mode.")

    captcha_dir = Path(args.captcha_dir).expanduser().resolve() if args.captcha_dir else workbook_path.parent / ".captchas"
    session = create_session()
    bteb_session = create_bteb_session()

    print(f"Result Data V2 version {VERSION}")
    print(f"Using workbook: {workbook_path}")
    print("Starting data extraction with the education board v2 result service.")
    print("Rows with any existing Name value are skipped. Clear the Name cell to retry a row.")
    if not args.no_browser_verify:
        print("Browser verification bootstrap is enabled for the education-board challenge page.")
    if not args.no_enter_pause:
        print("Press Enter during processing to save progress and pause at the next safe checkpoint.")

    for row_idx in range(2, sheet.max_row + 1):
        values = {
            "exam": excel_text(sheet.cell(row=row_idx, column=column_map["Exam"]).value),
            "year": excel_text(sheet.cell(row=row_idx, column=column_map["Passing Year"]).value),
            "board": excel_text(sheet.cell(row=row_idx, column=column_map["Board"]).value),
            "roll": excel_text(sheet.cell(row=row_idx, column=column_map["Roll"]).value),
            "reg": excel_text(sheet.cell(row=row_idx, column=column_map["Reg"]).value),
        }

        if all(not value for value in values.values()):
            continue

        if not pause_if_enter_requested(args, workbook, workbook_path, sheet, column_map):
            return 1

        print(f"\n--- Processing Row {row_idx} ---")

        if should_skip_row(sheet, row_idx, column_map, args.force):
            current_name = sheet.cell(row=row_idx, column=column_map["Name"]).value
            print(f"Skipping row {row_idx}: Name already present ('{current_name}'). Clear it to retry.")
            continue

        missing = [name for name, value in values.items() if not value]
        if missing:
            print(f"Skipping row {row_idx}: Missing input data: {', '.join(missing)}")
            write_error(sheet, row_idx, column_map, "Missing Input Data")
            if not save_progress_and_handle_pause(args, workbook, workbook_path, sheet, column_map):
                return 1
            continue

        mapped_exam = EXAM_TYPE_MAPPING.get(normalize_key(values["exam"]))
        mapped_board = BOARD_MAPPING.get(normalize_key(values["board"]))

        if not mapped_board:
            print(f"Skipping row {row_idx}: Unknown Board '{values['board']}'.")
            write_error(sheet, row_idx, column_map, "Unknown Board")
            if not save_progress_and_handle_pause(args, workbook, workbook_path, sheet, column_map):
                return 1
            continue

        try:
            if is_technical_board(values["board"], mapped_board):
                curriculum = infer_bteb_curriculum(
                    values["exam"],
                    get_optional_cell(sheet, row_idx, column_map, "curriculum"),
                )
                semester = infer_bteb_semester(
                    values["exam"],
                    get_optional_cell(sheet, row_idx, column_map, "semester"),
                )
                print(
                    "Using BTEB technical result service "
                    f"(curriculum={curriculum}, semester={semester})."
                )
                result = submit_bteb_result_request(
                    bteb_session,
                    curriculum_code=curriculum,
                    semester=semester,
                    year=values["year"],
                    roll=values["roll"],
                    reg=values["reg"],
                )
            else:
                if not mapped_exam:
                    print(f"Skipping row {row_idx}: Unknown Exam Type '{values['exam']}'.")
                    write_error(sheet, row_idx, column_map, "Unknown Exam Type")
                    if not save_progress_and_handle_pause(args, workbook, workbook_path, sheet, column_map):
                        return 1
                    continue

                result = lookup_result(
                    session,
                    board=mapped_board,
                    exam=mapped_exam,
                    year=values["year"],
                    roll=values["roll"],
                    reg=values["reg"],
                    row_idx=row_idx,
                    captcha_dir=captcha_dir,
                    open_captcha=not args.no_open_captcha,
                    max_attempts=args.max_captcha_attempts,
                    use_api=use_api,
                    api_key=api_key,
                    allow_browser_verify=not args.no_browser_verify,
                    browser_verify_timeout=args.browser_verify_timeout,
                    headless_browser_verify=args.headless_browser_verify,
                )
        except requests.RequestException as exc:
            print(f"HTTP error for row {row_idx}: {exc}")
            write_error(sheet, row_idx, column_map, "HTTP Error")
            if not save_progress_and_handle_pause(args, workbook, workbook_path, sheet, column_map):
                return 1
            continue
        except LookupErrorWithStatus as exc:
            print(f"{exc.status} for row {row_idx}: {exc.message}")
            write_error(sheet, row_idx, column_map, exc.status)
            if not save_progress_and_handle_pause(args, workbook, workbook_path, sheet, column_map):
                return 1
            continue
        except KeyboardInterrupt:
            print("\nStopped by user. Saving progress before exiting.")
            if not save_workbook_progress(workbook, workbook_path, sheet, column_map):
                return 1
            break
        except Exception as exc:
            print(f"Unexpected error for row {row_idx}: {exc}")
            write_error(sheet, row_idx, column_map, "Extraction Error")
            if not save_progress_and_handle_pause(args, workbook, workbook_path, sheet, column_map):
                return 1
            continue

        write_result(sheet, row_idx, column_map, result)
        print(
            "Successfully extracted: "
            f"Name='{result.name}', Father='{result.father_name}', GPA='{result.gpa}', "
            f"DoB='{result.date_of_birth}'"
        )
        if not save_progress_and_handle_pause(args, workbook, workbook_path, sheet, column_map):
            return 1
        time.sleep(args.delay)

    if not save_workbook_progress(workbook, workbook_path, sheet, column_map):
        return 1

    print(f"\nData extraction complete. Results saved to '{workbook_path}'")
    return 0


def pause_before_exit(enabled: bool) -> None:
    if not enabled:
        return
    try:
        input("\nFinished. Press Enter to close this window...")
    except EOFError:
        pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Fetch Bangladesh Education Board results from the current v2 website "
            f"and write them into {WORKBOOK_NAME}."
        )
    )
    parser.add_argument("--version", action="version", version=f"%(prog)s {VERSION}")
    parser.add_argument(
        "--file",
        help=(
            f"Workbook path. Defaults to '{WORKBOOK_NAME}' in the current working directory, "
            "then beside this script."
        ),
    )
    parser.add_argument(
        "--gemini-api-key",
        help="Gemini API key to automatically solve CAPTCHAs. Can also be set via the GEMINI_API_KEY environment variable.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Reprocess rows even when the Name column already contains a value.",
    )
    parser.add_argument(
        "--captcha-dir",
        help="Folder where CAPTCHA images should be saved. Defaults to a .captchas folder beside the workbook.",
    )
    parser.add_argument(
        "--no-open-captcha",
        action="store_true",
        help="Save CAPTCHA images without opening them automatically.",
    )
    parser.add_argument(
        "--no-browser-verify",
        action="store_true",
        help="Disable the browser verification bootstrap for the education-board challenge page.",
    )
    parser.add_argument(
        "--browser-verify-timeout",
        type=int,
        default=45,
        help="Seconds to wait for the education-board browser verification page. Default: 45.",
    )
    parser.add_argument(
        "--headless-browser-verify",
        action="store_true",
        help="Run the browser verification bootstrap without showing a browser window.",
    )
    parser.add_argument(
        "--no-pause",
        action="store_true",
        help="Exit immediately after finishing instead of waiting for Enter.",
    )
    parser.add_argument(
        "--no-enter-pause",
        action="store_true",
        help="Disable the Enter key pause/resume checkpoint while processing rows.",
    )
    parser.add_argument(
        "--max-captcha-attempts",
        type=int,
        default=3,
        help="Maximum CAPTCHA attempts per row. Default: 3.",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=2.0,
        help="Seconds to wait between successful result lookups. Default: 2.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    exit_code = extract_results_to_excel(args)
    pause_before_exit(not args.no_pause)
    return exit_code


if __name__ == "__main__":
    sys.exit(main())
