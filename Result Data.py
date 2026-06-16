from __future__ import annotations

import argparse
import os
import re
import sys
import time
import webbrowser
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import requests
from bs4 import BeautifulSoup


WORKBOOK_NAME = "Result data.xlsx"

BASE_URL = "https://www.educationboardresults.gov.bd"
HOME_URL = f"{BASE_URL}/v2/home"
CAPTCHA_URL = f"{BASE_URL}/v2/captcha"
RESULT_URL = f"{BASE_URL}/v2/getres"

INPUT_COLUMNS = ["Exam", "Passing Year", "Board", "Roll", "Reg"]
OUTPUT_COLUMNS = ["Name", "Father's Name", "Mother's Name", "GPA", "Date of Birth"]

ERROR_STATUSES = {
    "Result Not Found",
    "Invalid Criteria",
    "HTTP Error",
    "Extraction Error",
    "Missing Input Data",
    "Captcha Not Found",
    "Captcha Parse Error",
    "Unknown Exam Type",
    "Unknown Board",
    "HTML Parse Error (Captcha)",
    "Unexpected Error during Captcha",
    "Extraction Failed",
    "Manual Captcha Skipped",
    "Captcha Failed",
    "Result Error",
}

RAW_EXAM_TYPE_MAPPING = {
    "JSC/JDC": "jsc",
    "JSC or Equivalent": "jsc",
    "SSC/Dakhil/Equivalent": "ssc",
    "SSC/Dakhil": "ssc",
    "SSC or Equivalent": "ssc",
    "SSC(Vocational)": "ssc",
    "HSC/Alim/Equivalent": "hsc",
    "HSC/Alim": "hsc",
    "HSC(Vocational)": "hsc",
    "HSC(BM)": "hsc",
    "Diploma in Commerce": "hsc",
    "DIBS (Diploma in Business Studies)": "dibs",
    "Diploma in Business Studies": "dibs",
    "DIBS": "dibs",
}

RAW_BOARD_MAPPING = {
    "Barisal": "barisal",
    "Chittagong": "chittagong",
    "Comilla": "comilla",
    "Dhaka": "dhaka",
    "Dinajpur": "dinajpur",
    "Jessore": "jessore",
    "Mymensingh": "mymensingh",
    "Rajshahi": "rajshahi",
    "Sylhet": "sylhet",
    "Madrasah": "madrasah",
    "Technical": "tec",
    "TEC": "tec",
}


@dataclass
class StudentResult:
    name: str
    father_name: str
    mother_name: str
    gpa: str
    date_of_birth: str


class LookupErrorWithStatus(Exception):
    def __init__(self, status: str, message: str):
        super().__init__(message)
        self.status = status
        self.message = message


def normalize_key(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


EXAM_TYPE_MAPPING = {normalize_key(key): value for key, value in RAW_EXAM_TYPE_MAPPING.items()}
BOARD_MAPPING = {normalize_key(key): value for key, value in RAW_BOARD_MAPPING.items()}


def excel_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def resolve_workbook_path(path_arg: str | None) -> Path:
    if path_arg:
        return Path(path_arg).expanduser().resolve()

    cwd_candidate = Path.cwd() / WORKBOOK_NAME
    if cwd_candidate.exists():
        return cwd_candidate.resolve()

    script_candidate = Path(__file__).resolve().parent / WORKBOOK_NAME
    if script_candidate.exists():
        return script_candidate.resolve()

    return cwd_candidate.resolve()


def build_column_map(sheet: Any) -> dict[str, int]:
    column_map: dict[str, int] = {}
    for index, cell in enumerate(sheet[1], start=1):
        header = excel_text(cell.value)
        if header:
            column_map[header] = index
    return column_map


def require_columns(column_map: dict[str, int]) -> None:
    missing = [column for column in INPUT_COLUMNS + OUTPUT_COLUMNS if column not in column_map]
    if missing:
        joined = ", ".join(missing)
        raise RuntimeError(f"Required column(s) not found in the Excel sheet: {joined}")


def create_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "en-US,en;q=0.9",
            "Referer": HOME_URL,
        }
    )
    return session


def open_captcha_image(path: Path) -> None:
    try:
        if os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
        else:
            webbrowser.open(path.as_uri())
    except Exception as exc:
        print(f"Could not open CAPTCHA image automatically: {exc}")


def fetch_captcha(session: requests.Session, captcha_dir: Path, row_idx: int) -> Path:
    session.get(HOME_URL, timeout=60).raise_for_status()

    response = session.get(
        CAPTCHA_URL,
        params={"t": str(int(time.time() * 1000))},
        headers={"Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8"},
        timeout=60,
    )
    response.raise_for_status()

    content_type = response.headers.get("Content-Type", "")
    if "image" not in content_type.lower():
        raise LookupErrorWithStatus("Captcha Not Found", "The CAPTCHA endpoint did not return an image.")

    captcha_dir.mkdir(parents=True, exist_ok=True)
    captcha_path = captcha_dir / f"captcha_row_{row_idx}.png"
    captcha_path.write_bytes(response.content)
    return captcha_path


def prompt_for_captcha(row_idx: int, captcha_path: Path, should_open: bool) -> str:
    print(f"CAPTCHA for row {row_idx} saved to: {captcha_path}")
    if should_open:
        open_captcha_image(captcha_path)

    captcha = input(f"Enter CAPTCHA digits for row {row_idx} (blank skips this row): ").strip()
    if not captcha:
        raise LookupErrorWithStatus("Manual Captcha Skipped", "No CAPTCHA value entered.")
    return captcha


def extract_gpa(value: Any) -> str:
    text = excel_text(value)
    if not text:
        return "N/A"

    match = re.search(r"GPA\s*=?\s*([0-9.]+)", text, flags=re.IGNORECASE)
    if match:
        return match.group(1)
    return text


def parse_json_result(payload: dict[str, Any]) -> StudentResult:
    status = str(payload.get("status", ""))
    if status != "0":
        message = excel_text(payload.get("msg")) or "Result lookup failed."
        if "captcha" in message.lower():
            raise LookupErrorWithStatus("Captcha Failed", message)
        raise LookupErrorWithStatus("Result Error", message)

    result = payload.get("res")
    if not isinstance(result, dict):
        raise LookupErrorWithStatus("Extraction Failed", "The response did not contain a result object.")

    return StudentResult(
        name=excel_text(result.get("name")) or "N/A",
        father_name=excel_text(result.get("fname")) or "N/A",
        mother_name=excel_text(result.get("mname")) or "N/A",
        gpa=extract_gpa(result.get("gpa") or result.get("res_detail") or result.get("result")),
        date_of_birth=excel_text(result.get("dob")) or "N/A",
    )


def table_pairs_from_html(html: str) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    pairs: dict[str, str] = {}

    for row in soup.select("tr"):
        cells = [cell.get_text(" ", strip=True) for cell in row.find_all(["td", "th"])]
        if len(cells) >= 2:
            pairs[normalize_key(cells[0])] = cells[1]
        if len(cells) >= 4:
            pairs[normalize_key(cells[2])] = cells[3]

    return pairs


def parse_html_result(html: str) -> StudentResult:
    pairs = table_pairs_from_html(html)
    if not pairs:
        raise LookupErrorWithStatus("Extraction Failed", "No result table was found in the HTML response.")

    name = pairs.get("name of student") or pairs.get("name") or "N/A"
    father_name = pairs.get("father's name") or "N/A"
    mother_name = pairs.get("mother's name") or "N/A"
    gpa = pairs.get("gpa") or extract_gpa(pairs.get("result"))
    date_of_birth = pairs.get("date of birth") or "N/A"

    if all(value == "N/A" for value in [name, father_name, mother_name, gpa, date_of_birth]):
        raise LookupErrorWithStatus("Extraction Failed", "The result table did not contain expected fields.")

    return StudentResult(
        name=name,
        father_name=father_name,
        mother_name=mother_name,
        gpa=gpa,
        date_of_birth=date_of_birth,
    )


def submit_result_request(
    session: requests.Session,
    *,
    board: str,
    exam: str,
    year: str,
    roll: str,
    reg: str,
    captcha: str,
) -> StudentResult:
    payload = {
        "board": board,
        "exam": exam,
        "year": year,
        "result_type": "1",
        "roll": roll,
        "reg": reg,
        "captcha": captcha,
    }

    response = session.post(
        RESULT_URL,
        data=payload,
        headers={
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Origin": BASE_URL,
            "Referer": HOME_URL,
            "X-Requested-With": "XMLHttpRequest",
        },
        timeout=60,
    )
    response.raise_for_status()

    try:
        return parse_json_result(response.json())
    except ValueError:
        return parse_html_result(response.text)


def lookup_result_with_manual_captcha(
    session: requests.Session,
    *,
    board: str,
    exam: str,
    year: str,
    roll: str,
    reg: str,
    row_idx: int,
    captcha_dir: Path,
    open_captcha: bool,
    max_attempts: int,
) -> StudentResult:
    last_error: LookupErrorWithStatus | None = None

    for attempt in range(1, max_attempts + 1):
        captcha_path = fetch_captcha(session, captcha_dir, row_idx)
        captcha = prompt_for_captcha(row_idx, captcha_path, open_captcha)

        try:
            return submit_result_request(
                session,
                board=board,
                exam=exam,
                year=year,
                roll=roll,
                reg=reg,
                captcha=captcha,
            )
        except LookupErrorWithStatus as exc:
            last_error = exc
            print(f"Attempt {attempt}/{max_attempts} failed for row {row_idx}: {exc.message}")
            if exc.status != "Captcha Failed":
                raise

    if last_error:
        raise last_error
    raise LookupErrorWithStatus("Captcha Failed", "CAPTCHA attempts were exhausted.")


def write_result(sheet: Any, row_idx: int, column_map: dict[str, int], result: StudentResult) -> None:
    sheet.cell(row=row_idx, column=column_map["Name"], value=result.name)
    sheet.cell(row=row_idx, column=column_map["Father's Name"], value=result.father_name)
    sheet.cell(row=row_idx, column=column_map["Mother's Name"], value=result.mother_name)
    sheet.cell(row=row_idx, column=column_map["GPA"], value=result.gpa)
    sheet.cell(row=row_idx, column=column_map["Date of Birth"], value=result.date_of_birth)


def write_error(sheet: Any, row_idx: int, column_map: dict[str, int], status: str) -> None:
    sheet.cell(row=row_idx, column=column_map["Name"], value=status)
    for column in OUTPUT_COLUMNS[1:]:
        sheet.cell(row=row_idx, column=column_map[column], value=None)


def should_skip_row(sheet: Any, row_idx: int, column_map: dict[str, int], force: bool) -> bool:
    if force:
        return False

    current_name = excel_text(sheet.cell(row=row_idx, column=column_map["Name"]).value)
    return bool(current_name and current_name not in ERROR_STATUSES)


def extract_results_to_excel(args: argparse.Namespace) -> int:
    workbook_path = resolve_workbook_path(args.file)
    if not workbook_path.exists():
        print(f"Error: Excel file not found at '{workbook_path}'")
        return 1

    try:
        workbook = openpyxl.load_workbook(workbook_path)
        sheet = workbook.active
    except Exception as exc:
        print(f"Error loading Excel file: {exc}")
        return 1

    try:
        column_map = build_column_map(sheet)
        require_columns(column_map)
    except RuntimeError as exc:
        print(f"Error: {exc}")
        return 1

    captcha_dir = Path(args.captcha_dir).expanduser().resolve() if args.captcha_dir else workbook_path.parent / ".captchas"
    session = create_session()

    print(f"Using workbook: {workbook_path}")
    print("Starting data extraction with the education board v2 result service.")

    for row_idx in range(2, sheet.max_row + 1):
        values = {
            "exam": excel_text(sheet.cell(row=row_idx, column=column_map["Exam"]).value),
            "year": excel_text(sheet.cell(row=row_idx, column=column_map["Passing Year"]).value),
            "board": excel_text(sheet.cell(row=row_idx, column=column_map["Board"]).value),
            "roll": excel_text(sheet.cell(row=row_idx, column=column_map["Roll"]).value),
            "reg": excel_text(sheet.cell(row=row_idx, column=column_map["Reg"]).value),
        }

        if all(not value for value in values.values()):
            continue

        print(f"\n--- Processing Row {row_idx} ---")

        if should_skip_row(sheet, row_idx, column_map, args.force):
            current_name = sheet.cell(row=row_idx, column=column_map["Name"]).value
            print(f"Skipping row {row_idx}: Name already present ('{current_name}').")
            continue

        missing = [name for name, value in values.items() if not value]
        if missing:
            print(f"Skipping row {row_idx}: Missing input data: {', '.join(missing)}")
            write_error(sheet, row_idx, column_map, "Missing Input Data")
            continue

        mapped_exam = EXAM_TYPE_MAPPING.get(normalize_key(values["exam"]))
        mapped_board = BOARD_MAPPING.get(normalize_key(values["board"]))

        if not mapped_exam:
            print(f"Skipping row {row_idx}: Unknown Exam Type '{values['exam']}'.")
            write_error(sheet, row_idx, column_map, "Unknown Exam Type")
            continue

        if not mapped_board:
            print(f"Skipping row {row_idx}: Unknown Board '{values['board']}'.")
            write_error(sheet, row_idx, column_map, "Unknown Board")
            continue

        try:
            result = lookup_result_with_manual_captcha(
                session,
                board=mapped_board,
                exam=mapped_exam,
                year=values["year"],
                roll=values["roll"],
                reg=values["reg"],
                row_idx=row_idx,
                captcha_dir=captcha_dir,
                open_captcha=not args.no_open_captcha,
                max_attempts=args.max_captcha_attempts,
            )
        except requests.RequestException as exc:
            print(f"HTTP error for row {row_idx}: {exc}")
            write_error(sheet, row_idx, column_map, "HTTP Error")
            continue
        except LookupErrorWithStatus as exc:
            print(f"{exc.status} for row {row_idx}: {exc.message}")
            write_error(sheet, row_idx, column_map, exc.status)
            continue
        except KeyboardInterrupt:
            print("\nStopped by user. Saving progress before exiting.")
            break
        except Exception as exc:
            print(f"Unexpected error for row {row_idx}: {exc}")
            write_error(sheet, row_idx, column_map, "Extraction Error")
            continue

        write_result(sheet, row_idx, column_map, result)
        print(
            "Successfully extracted: "
            f"Name='{result.name}', Father='{result.father_name}', GPA='{result.gpa}', "
            f"DoB='{result.date_of_birth}'"
        )
        time.sleep(args.delay)

    try:
        workbook.save(workbook_path)
    except PermissionError:
        print(f"\nError saving Excel file. Close '{workbook_path.name}' if it is open, then run again.")
        return 1
    except Exception as exc:
        print(f"\nError saving Excel file: {exc}")
        return 1

    print(f"\nData extraction complete. Results saved to '{workbook_path}'")
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Fetch Bangladesh Education Board results from the current v2 website "
            "and write them into Result data.xlsx."
        )
    )
    parser.add_argument(
        "--file",
        help=(
            "Workbook path. Defaults to 'Result data.xlsx' in the current working directory, "
            "then beside this script."
        ),
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Reprocess rows even when the Name column already contains a successful value.",
    )
    parser.add_argument(
        "--captcha-dir",
        help="Folder where CAPTCHA images should be saved. Defaults to a .captchas folder beside the workbook.",
    )
    parser.add_argument(
        "--no-open-captcha",
        action="store_true",
        help="Save CAPTCHA images without opening them automatically.",
    )
    parser.add_argument(
        "--max-captcha-attempts",
        type=int,
        default=3,
        help="Maximum CAPTCHA attempts per row. Default: 3.",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=2.0,
        help="Seconds to wait between successful result lookups. Default: 2.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    sys.exit(extract_results_to_excel(parse_args()))
